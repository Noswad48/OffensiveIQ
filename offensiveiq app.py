import streamlit as st
import pandas as pd
import io
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl

st.set_page_config(
    page_title="OffensiveIQ — Defensive Tendency Analysis",
    page_icon="🏈",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@700;800;900&family=Barlow:wght@400;500;600&family=Share+Tech+Mono&display=swap');
html,body,[class*="css"]{font-family:'Barlow',sans-serif;background-color:#0a1628;color:#f0ede8;}
.stApp{background-color:#0a1628;}
.main-title{font-family:'Barlow Condensed',sans-serif;font-weight:900;font-size:64px;line-height:.95;text-transform:uppercase;color:#f0ede8;margin-bottom:8px;}
.stButton>button{background:#1a5276!important;color:#f0ede8!important;border:none!important;font-family:'Barlow Condensed',sans-serif!important;font-weight:700!important;font-size:16px!important;letter-spacing:.1em!important;text-transform:uppercase!important;padding:12px 32px!important;border-radius:0!important;width:100%!important;}
.stButton>button:hover{background:#154360!important;}
.stDownloadButton>button{background:#0e7060!important;color:#f0ede8!important;border:none!important;font-family:'Barlow Condensed',sans-serif!important;font-weight:700!important;font-size:14px!important;letter-spacing:.08em!important;text-transform:uppercase!important;border-radius:0!important;width:100%!important;}
</style>
""", unsafe_allow_html=True)

# ── Helpers ───────────────────────────────────────────────────
def get_zone(y):
    try:
        y=float(y)
        if y<=-1  and y>=-20: return "BZ"
        if y<=-21 and y>=-49: return "OF"
        if y>=40  and y<=50:  return "MF"
        if y>=21  and y<=39:  return "FZ"
        if y>=11  and y<=20:  return "RZ"
        if y>=1   and y<=10:  return "GL"
    except: pass
    return None

def top3(plays, key):
    vals=[str(p.get(key,'')) for p in plays
          if p.get(key) not in (None,'') and str(p.get(key,'')).strip() not in ('','nan','None')]
    if not vals: return []
    return [{"v":v,"n":c} for v,c in Counter(vals).most_common(3)]

def pct(n,d): return round(n/d*100) if d>0 else 0

def load_plays(df):
    plays=[]
    for _,row in df.iterrows():
        pt=str(row.get('PLAY TYPE','')).strip()
        if pt not in ('Run','Pass'): continue
        zone=get_zone(row.get('YARD LN',''))
        if not zone: continue
        plays.append({
            'zone':  zone,
            'dn':    int(float(row.get('DN',0) or 0)),
            'dist':  float(row.get('DIST',0) or 0),
            'hash':  str(row.get('HASH','')).strip(),
            'form':  str(row.get('OFF FORM','')).strip(),
            'play':  str(row.get('OFF PLAY','')).strip(),
            'dir':   str(row.get('PLAY DIR','')).strip(),
            'rp':    pt,
            'gnls':  float(row.get('GN/LS',0) or 0),
            'front': str(row.get('DEF FRONT','')).strip(),
            'cov':   str(row.get('COVERAGE','')).strip(),
            'blitz': str(row.get('BLITZ','')).strip(),
            'result':str(row.get('RESULT','')).strip(),
        })
    return plays

# ── Excel Builder ─────────────────────────────────────────────
def build_excel(plays, opp, week, date):
    FN="Arial"; CW="FFFFFFFF"; CL="FFF5F5F5"; CB="FF16213E"
    CBl="FF1A5276"; CTe="FF0E7060"; CPu="FF4A235A"; CR="FFC0392B"
    CDG="FF555555"; CGr="FF1E8449"
    # Zone colors — blue theme for offensive version
    ZONE_BG={"BZ":"FFFDE8E8","OF":"FFE8F0FE","MF":"FFE8F8E8",
              "FZ":"FFFFFBE6","RZ":"FFFCE4EC","GL":"FFEDE7F6"}
    ZONE_HDR={"BZ":CR,"OF":CBl,"MF":CTe,"FZ":"FF7D6608","RZ":CR,"GL":CPu}

    def fil(c): return PatternFill("solid",fgColor=c)
    def bdr():
        s=Side(style="thin",color="FFB0B0B0")
        return Border(left=s,right=s,top=s,bottom=s)
    def sc(ws,r,c,val="",bold=False,sz=10,fc=CB,bg=None,h="center",v="center",wrap=False,fmt=None):
        cell=ws.cell(row=r,column=c,value=val)
        cell.font=Font(name=FN,bold=bold,size=sz,color=fc)
        if bg: cell.fill=fil(bg)
        cell.alignment=Alignment(horizontal=h,vertical=v,wrap_text=wrap)
        cell.border=bdr()
        if fmt: cell.number_format=fmt
        return cell
    def hdr(ws,r,c,txt,bg=CBl,fc=CW,sz=9,wrap=True,span=1):
        cell=sc(ws,r,c,txt,bold=True,sz=sz,fc=fc,bg=bg,wrap=wrap)
        if span>1: ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+span-1)
        return cell
    def banner(ws,r,txt,nc,bg=CB,fc=CW,sz=13,ht=30):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=nc)
        c=ws.cell(row=r,column=1,value=txt)
        c.font=Font(name=FN,bold=True,size=sz,color=fc)
        c.fill=fil(bg); c.alignment=Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[r].height=ht
    def widths(ws,lst):
        for i,w in enumerate(lst,1): ws.column_dimensions[gcl(i)].width=w

    zone_list=["BZ","OF","MF","FZ","RZ","GL"]
    zone_names={"BZ":"Backed Up  Own 1–20","OF":"Open Field  Own 21–49",
                "MF":"Midfield  50–Opp 40","FZ":"Fringe  Opp 39–21",
                "RZ":"Red Zone  Opp 20–11","GL":"Goal Line  Opp 10 and in"}
    runs=[p for p in plays if p['rp']=='Run']
    passes=[p for p in plays if p['rp']=='Pass']
    total=len(plays)
    has_front=any(p['front'] for p in plays)
    has_cov  =any(p['cov']   for p in plays)
    has_blitz=any(p['blitz'] for p in plays)

    dd=[
        ("1st Down",  lambda p: p['dn']==1),
        ("2nd & 7+",  lambda p: p['dn']==2 and p['dist']>=7),
        ("2nd & 4-6", lambda p: p['dn']==2 and 4<=p['dist']<=6),
        ("2nd & 1-3", lambda p: p['dn']==2 and p['dist']<=3),
        ("3rd & 7+",  lambda p: p['dn']==3 and p['dist']>=7),
        ("3rd & 4-6", lambda p: p['dn']==3 and 4<=p['dist']<=6),
        ("3rd & 1-3", lambda p: p['dn']==3 and p['dist']<=3),
        ("4th Down",  lambda p: p['dn']==4),
    ]

    wb2=Workbook()

    # ── Tab 1: Film Log ──────────────────────────────────────
    ws_log=wb2.active; ws_log.title="1. Film Log"
    ws_log.sheet_properties.tabColor="1A5276"
    ws_log.sheet_view.showGridLines=False
    log_cols=[('QTR',6),('DN',6),('DIST',6),('HASH',6),('YARD LN',9),('ZONE',10),
              ('OFF FORM',18),('OFF PLAY',20),('PLAY DIR',9),('PLAY TYPE',10),
              ('GN/LS',8),('RESULT',10),('DEF FRONT',14),('COVERAGE',14),('BLITZ',14)]
    widths(ws_log,[w for _,w in log_cols])
    banner(ws_log,1,"OFFENSIVE IQ — FILM LOG  |  Tag DEF FRONT, COVERAGE, BLITZ in Hudl for full analysis",
           len(log_cols),bg=CBl,sz=10,ht=26)
    ws_log.row_dimensions[2].height=36
    for ci,(col,_) in enumerate(log_cols,1):
        if col=="ZONE":    bg="FF000088"
        elif col in("DEF FRONT","COVERAGE","BLITZ"): bg=CR
        else: bg=CBl
        hdr(ws_log,2,ci,col,bg=bg,sz=8)
    for ri,p in enumerate(plays):
        r=ri+3; ws_log.row_dimensions[r].height=15
        bg=CL if ri%2==0 else CW
        vals={'QTR':'','DN':p['dn'],'DIST':p['dist'],'HASH':p['hash'],
              'YARD LN':'','ZONE':p['zone'],'OFF FORM':p['form'],'OFF PLAY':p['play'],
              'PLAY DIR':p['dir'],'PLAY TYPE':p['rp'],'GN/LS':p['gnls'],
              'RESULT':p['result'],'DEF FRONT':p['front'],'COVERAGE':p['cov'],'BLITZ':p['blitz']}
        for ci,(col,_) in enumerate(log_cols,1):
            zbg="FFE8F4FD" if col=="ZONE" else ("FFFDE8E8" if col in("DEF FRONT","COVERAGE","BLITZ") else bg)
            sc(ws_log,r,ci,vals.get(col,''),sz=9,bg=zbg,fc="FF000000",
               h="left" if col in("OFF FORM","OFF PLAY","RESULT","DEF FRONT","COVERAGE","BLITZ") else "center")
    ws_log.freeze_panes="A3"

    # ── Tab 2: Field Zone Tendencies ─────────────────────────
    ws2=wb2.create_sheet("2. Field Zone Tendencies")
    ws2.sheet_properties.tabColor="1A5276"; ws2.sheet_view.showGridLines=False
    NC2=2+len(dd); widths(ws2,[18,14]+[18]*len(dd))
    banner(ws2,1,"FIELD ZONE TENDENCIES — What defense faces us where",NC2,bg=CB,sz=13,ht=32)
    ws2.merge_cells(f"A2:{gcl(NC2)}2")
    leg=ws2.cell(row=2,column=1,value="  Gray=plays  Red%=Run tendency  Blue%=Pass tendency  Yellow=OC note")
    leg.font=Font(name=FN,size=8,italic=True,color=CDG)
    leg.fill=fil("FFF0F0F0"); leg.alignment=Alignment(horizontal="left",vertical="center")
    ws2.row_dimensions[2].height=14
    ws2.row_dimensions[3].height=36
    sc(ws2,3,1,"FIELD ZONE",bold=True,sz=9,fc=CW,bg=CB)
    sc(ws2,3,2,"METRIC",    bold=True,sz=9,fc=CW,bg=CB)
    for ci,(lbl,_) in enumerate(dd): hdr(ws2,3,ci+3,lbl,bg=CB,sz=9)

    sub=[("Plays","count"),("Run %","runpct"),("Pass %","passpct"),
         ("Runs","runcnt"),("Passes","passcnt"),("▶ OC Note","call")]
    row=4
    for zcode in zone_list:
        zbg=ZONE_BG[zcode]; zhdr=ZONE_HDR[zcode]
        zplays=[p for p in plays if p['zone']==zcode]
        ws2.row_dimensions[row].height=17
        ws2.merge_cells(start_row=row,start_column=1,end_row=row,end_column=NC2)
        c=ws2.cell(row=row,column=1,value=f"  {zcode}  ·  {zone_names[zcode]}  ({len(zplays)} plays)")
        c.font=Font(name=FN,bold=True,size=10,color=CW)
        c.fill=fil(zhdr); c.alignment=Alignment(horizontal="left",vertical="center")
        row+=1
        for slbl,stype in sub:
            ws2.row_dimensions[row].height=16
            sc(ws2,row,1,"",bg=zbg)
            sc(ws2,row,2,slbl,bold=(stype in("runpct","passpct")),sz=9,fc=CDG,bg=zbg,h="right")
            for ci_i,(_,dfilt) in enumerate(dd):
                cn=ci_i+3; c=ws2.cell(row=row,column=cn)
                c.border=bdr(); c.alignment=Alignment(horizontal="center",vertical="center")
                try: filtered=[p for p in zplays if dfilt(p)]
                except: filtered=[]
                n_run=len([p for p in filtered if p['rp']=='Run'])
                n_pass=len([p for p in filtered if p['rp']=='Pass'])
                tot=len(filtered)
                if stype=="count":
                    c.value=tot; c.fill=fil("FFE8E8E8"); c.font=Font(name=FN,sz=9,color="FF000000"); c.number_format="0"
                elif stype=="runpct":
                    c.value=round(n_run/tot,2) if tot>0 else ""; c.font=Font(name=FN,bold=True,sz=11,color="FF8B0000"); c.fill=fil(zbg); c.number_format="0%"
                elif stype=="passpct":
                    c.value=round(n_pass/tot,2) if tot>0 else ""; c.font=Font(name=FN,bold=True,sz=11,color="FF00008B"); c.fill=fil(zbg); c.number_format="0%"
                elif stype=="runcnt":
                    c.value=n_run; c.fill=fil("FFFDE8E8"); c.font=Font(name=FN,sz=9,color="FF8B0000"); c.number_format="0"
                elif stype=="passcnt":
                    c.value=n_pass; c.fill=fil("FFE8F0FE"); c.font=Font(name=FN,sz=9,color="FF00008B"); c.number_format="0"
                elif stype=="call":
                    c.fill=fil("FFFFFBE6"); c.font=Font(name=FN,sz=8,italic=True,color=CDG)
                    c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
            row+=1
    ws2.freeze_panes="C4"

    # ── Tab 3: Defensive Fronts ──────────────────────────────
    ws3=wb2.create_sheet("3. Defensive Fronts")
    ws3.sheet_properties.tabColor="C0392B"; ws3.sheet_view.showGridLines=False
    NC3=10; widths(ws3,[8,12,9,10,20,20,20,20,20,28])
    banner(ws3,1,"DEFENSIVE FRONTS  —  What fronts do we face by zone",NC3,bg=CR,sz=12,ht=30)
    ws3.row_dimensions[2].height=20
    for c,txt,bg,span in[(1,"ZONE",CB,1),(2,"COUNTS",CB,3),(5,"TOP FRONTS FACED",CR,3),(8,"TOP FRONT vs RUN",CR,2),(10,"NOTES",CB,1)]:
        hdr(ws3,2,c,txt,bg=bg,sz=8,span=span)
    ws3.row_dimensions[3].height=36
    for c,txt,bg in[(1,"Zone",CB),(2,"Plays",CB),(3,"Run",CB),(4,"Pass",CB),
                     (5,"#1 Front",CR),(6,"#2 Front",CR),(7,"#3 Front",CR),
                     (8,"#1 vs Run",CR),(9,"#2 vs Run",CR),(10,"Notes",CB)]:
        hdr(ws3,3,c,txt,bg=bg,sz=9,wrap=True)
    for ri,zcode in enumerate(zone_list):
        r=ri+4; zbg=ZONE_BG[zcode]; zhdr=ZONE_HDR[zcode]
        zp=[p for p in plays if p['zone']==zcode]
        zr=[p for p in zp if p['rp']=='Run']
        zpass=[p for p in zp if p['rp']=='Pass']
        ws3.row_dimensions[r].height=28
        sc(ws3,r,1,zcode,bold=True,sz=11,fc=CW,bg=zhdr)
        sc(ws3,r,2,len(zp),bold=True,sz=11,fc="FF000000",bg=zbg,fmt="0")
        sc(ws3,r,3,len(zr),bold=True,sz=11,fc="FF8B0000",bg="FFFDE8E8",fmt="0")
        sc(ws3,r,4,len(zpass),bold=True,sz=11,fc="FF00008B",bg="FFE8F0FE",fmt="0")
        t3f=top3(zp,'front'); t3fr=top3(zr,'front')
        for i,cn in enumerate([5,6,7]): sc(ws3,r,cn,t3f[i]['v']+f" ({t3f[i]['n']})" if i<len(t3f) else "—",sz=9,bg=zbg,wrap=True)
        for i,cn in enumerate([8,9]): sc(ws3,r,cn,t3fr[i]['v']+f" ({t3fr[i]['n']})" if i<len(t3fr) else "—",sz=9,bg=zbg,wrap=True)
        sc(ws3,r,10,"",bg=zbg,sz=9,wrap=True,h="left")
    ws3.freeze_panes="B4"

    # ── Tab 4: Coverages ─────────────────────────────────────
    ws4=wb2.create_sheet("4. Coverages")
    ws4.sheet_properties.tabColor="1A5276"; ws4.sheet_view.showGridLines=False
    NC4=10; widths(ws4,[8,12,9,10,20,20,20,20,20,28])
    banner(ws4,1,"COVERAGES FACED  —  What coverage do we see by zone",NC4,bg=CBl,sz=12,ht=30)
    ws4.row_dimensions[2].height=20
    for c,txt,bg,span in[(1,"ZONE",CB,1),(2,"COUNTS",CB,3),(5,"TOP COVERAGES",CBl,3),(8,"TOP COV vs PASS",CBl,2),(10,"NOTES",CB,1)]:
        hdr(ws4,2,c,txt,bg=bg,sz=8,span=span)
    ws4.row_dimensions[3].height=36
    for c,txt,bg in[(1,"Zone",CB),(2,"Plays",CB),(3,"Run",CB),(4,"Pass",CB),
                     (5,"#1 Coverage",CBl),(6,"#2 Coverage",CBl),(7,"#3 Coverage",CBl),
                     (8,"#1 vs Pass",CBl),(9,"#2 vs Pass",CBl),(10,"Notes",CB)]:
        hdr(ws4,3,c,txt,bg=bg,sz=9,wrap=True)
    for ri,zcode in enumerate(zone_list):
        r=ri+4; zbg=ZONE_BG[zcode]; zhdr=ZONE_HDR[zcode]
        zp=[p for p in plays if p['zone']==zcode]
        zr=[p for p in zp if p['rp']=='Run']
        zpass=[p for p in zp if p['rp']=='Pass']
        ws4.row_dimensions[r].height=28
        sc(ws4,r,1,zcode,bold=True,sz=11,fc=CW,bg=zhdr)
        sc(ws4,r,2,len(zp),bold=True,sz=11,fc="FF000000",bg=zbg,fmt="0")
        sc(ws4,r,3,len(zr),bold=True,sz=11,fc="FF8B0000",bg="FFFDE8E8",fmt="0")
        sc(ws4,r,4,len(zpass),bold=True,sz=11,fc="FF00008B",bg="FFE8F0FE",fmt="0")
        t3c=top3(zp,'cov'); t3cp=top3(zpass,'cov')
        for i,cn in enumerate([5,6,7]): sc(ws4,r,cn,t3c[i]['v']+f" ({t3c[i]['n']})" if i<len(t3c) else "—",sz=9,bg=zbg,wrap=True)
        for i,cn in enumerate([8,9]): sc(ws4,r,cn,t3cp[i]['v']+f" ({t3cp[i]['n']})" if i<len(t3cp) else "—",sz=9,bg=zbg,wrap=True)
        sc(ws4,r,10,"",bg=zbg,sz=9,wrap=True,h="left")
    ws4.freeze_panes="B4"

    # ── Tab 5: Blitzes ───────────────────────────────────────
    ws5=wb2.create_sheet("5. Blitzes")
    ws5.sheet_properties.tabColor="4A235A"; ws5.sheet_view.showGridLines=False
    NC5=10; widths(ws5,[8,10,10,10,18,18,18,14,14,28])
    banner(ws5,1,"BLITZ TENDENCIES  —  When and where do we face pressure",NC5,bg=CPu,sz=12,ht=30)
    ws5.row_dimensions[2].height=20
    for c,txt,bg,span in[(1,"ZONE",CB,1),(2,"COUNTS",CB,3),(5,"TOP BLITZ TYPES",CPu,3),(8,"HASH BLITZ",CPu,2),(10,"NOTES",CB,1)]:
        hdr(ws5,2,c,txt,bg=bg,sz=8,span=span)
    ws5.row_dimensions[3].height=36
    for c,txt,bg in[(1,"Zone",CB),(2,"Plays",CB),(3,"Blitz\nCount",CB),(4,"Blitz %",CB),
                     (5,"#1 Blitz",CPu),(6,"#2 Blitz",CPu),(7,"#3 Blitz",CPu),
                     (8,"L Hash\nBlitz%",CPu),(9,"R Hash\nBlitz%",CPu),(10,"Notes",CB)]:
        hdr(ws5,3,c,txt,bg=bg,sz=9,wrap=True)
    for ri,zcode in enumerate(zone_list):
        r=ri+4; zbg=ZONE_BG[zcode]; zhdr=ZONE_HDR[zcode]
        zp=[p for p in plays if p['zone']==zcode]
        zb=[p for p in zp if p['blitz'] not in ('','nan','None','0','No')]
        zl=[p for p in zp if p['hash']=='L']; zr_h=[p for p in zp if p['hash']=='R']
        zbl=[p for p in zl if p['blitz'] not in ('','nan','None','0','No')]
        zbr=[p for p in zr_h if p['blitz'] not in ('','nan','None','0','No')]
        ws5.row_dimensions[r].height=28
        sc(ws5,r,1,zcode,bold=True,sz=11,fc=CW,bg=zhdr)
        sc(ws5,r,2,len(zp),bold=True,sz=11,fc="FF000000",bg=zbg,fmt="0")
        sc(ws5,r,3,len(zb),bold=True,sz=11,fc="FF4A235A",bg="FFEDE7F6",fmt="0")
        sc(ws5,r,4,round(len(zb)/len(zp),2) if zp else "",bold=True,sz=11,fc="FF4A235A",bg="FFEDE7F6",fmt="0%")
        t3b=top3(zb,'blitz')
        for i,cn in enumerate([5,6,7]): sc(ws5,r,cn,t3b[i]['v']+f" ({t3b[i]['n']})" if i<len(t3b) else "—",sz=9,bg=zbg,wrap=True)
        sc(ws5,r,8,round(len(zbl)/len(zl),2) if zl else "",sz=10,fc="FF6C3483",bg="FFEAF0FF",fmt="0%")
        sc(ws5,r,9,round(len(zbr)/len(zr_h),2) if zr_h else "",sz=10,fc="FF784212",bg="FFFFF0EA",fmt="0%")
        sc(ws5,r,10,"",bg=zbg,sz=9,wrap=True,h="left")
    ws5.freeze_panes="B4"

    # ── Tab 6: Hash Tendencies ───────────────────────────────
    ws6=wb2.create_sheet("6. Hash Tendencies")
    ws6.sheet_properties.tabColor="6C3483"; ws6.sheet_view.showGridLines=False
    NC6=13; widths(ws6,[18,10,10,10,10,10,10,18,18,18,18,18,18])
    banner(ws6,1,"HASH TENDENCIES  —  Left · Middle · Right",NC6,bg="FF6C3483",sz=13,ht=32)
    ws6.row_dimensions[2].height=36
    for cn,txt,bg in[(1,"FIELD ZONE",CB),(2,"L Plays","FF6C3483"),(3,"L Run%","FF6C3483"),(4,"L Pass%","FF6C3483"),
                      (5,"M Plays","FF1A5276"),(6,"M Run%","FF1A5276"),(7,"M Pass%","FF1A5276"),
                      (8,"R Plays","FF784212"),(9,"R Run%","FF784212"),(10,"R Pass%","FF784212"),
                      (11,"Top L Front",CB),(12,"Top M Front",CB),(13,"Top R Front",CB)]:
        hdr(ws6,2,cn,txt,bg=bg,sz=8)

    def hash_row(ws,r,label,base,zhdr_col,zbg):
        ws.row_dimensions[r].height=24
        sc(ws,r,1,label,bold=True,sz=10,fc=CW,bg=zhdr_col)
        for h,cols in[('L',(2,3,4)),('M',(5,6,7)),('R',(8,9,10))]:
            hp=[p for p in base if p['hash']==h]
            hr=[p for p in hp if p['rp']=='Run']
            hpass=[p for p in hp if p['rp']=='Pass']
            n=len(hp)
            sc(ws,r,cols[0],n,sz=11,bold=True,fc="FF000000",bg=zbg,fmt="0")
            sc(ws,r,cols[1],round(len(hr)/n,2) if n>0 else "",sz=11,bold=True,fc="FF8B0000",bg=zbg,fmt="0%")
            sc(ws,r,cols[2],round(len(hpass)/n,2) if n>0 else "",sz=11,bold=True,fc="FF00008B",bg=zbg,fmt="0%")
        for col_n,h in[(11,'L'),(12,'M'),(13,'R')]:
            hp=[p for p in base if p['hash']==h]
            tf=top3(hp,'front')
            sc(ws,r,col_n,tf[0]['v']+f" ({tf[0]['n']})" if tf else "—",sz=9,bg=zbg,h="left",wrap=True)

    hash_row(ws6,3,"OVERALL",plays,CB,CL)
    for ri,zcode in enumerate(zone_list):
        hash_row(ws6,ri+4,f"{zcode} — {zone_names[zcode]}",[p for p in plays if p['zone']==zcode],ZONE_HDR[zcode],ZONE_BG[zcode])
    ws6.freeze_panes="A3"

    # ── Tab 7: Situational Summary ───────────────────────────
    ws7=wb2.create_sheet("7. Situational Summary")
    ws7.sheet_properties.tabColor="4A235A"; ws7.sheet_view.showGridLines=False
    NC7=11; widths(ws7,[16,9,9,18,18,18,18,18,18,18,28])
    banner(ws7,1,"SITUATIONAL SUMMARY  —  What do we face in every situation",NC7,bg="FF4A235A",sz=12,ht=30)
    s7h=["Situation","Run\nCount","Pass\nCount","Top Front","Top Coverage",
         "Blitz %","L Hash\nRun%","M Hash\nRun%","R Hash\nRun%","Top Play","Notes"]
    ws7.row_dimensions[2].height=38
    for ci,h in enumerate(s7h): hdr(ws7,2,ci+1,h,bg="FF4A235A",sz=9,wrap=True)

    def sit(dn=None,dmin=None,dmax=None,zone=None):
        out=[]
        for p in plays:
            if dn   and p['dn']!=dn:     continue
            if dmin and p['dist']<dmin:  continue
            if dmax and p['dist']>dmax:  continue
            if zone and p['zone']!=zone: continue
            out.append(p)
        return out

    sits=[
        ("1ST DOWN",      dict(dn=1)),
        ("2ND & LONG",    dict(dn=2,dmin=7)),
        ("2ND & MEDIUM",  dict(dn=2,dmin=4,dmax=6)),
        ("2ND & SHORT",   dict(dn=2,dmax=3)),
        ("3RD & LONG",    dict(dn=3,dmin=7)),
        ("3RD & MEDIUM",  dict(dn=3,dmin=4,dmax=6)),
        ("3RD & SHORT",   dict(dn=3,dmax=3)),
        ("4TH DOWN",      dict(dn=4)),
        ("RED ZONE",      dict(zone="RZ")),
        ("GOAL LINE",     dict(zone="GL")),
        ("BACKED UP",     dict(zone="BZ")),
        ("COMING OUT",    dict(zone="OF")),
        ("TWO-MINUTE",    dict(dn=3,dmin=5)),
        ("MUST HAVE",     dict(dn=4)),
    ]
    sit_colors=["FF0E7060","FF1A5276","FF1A5276","FF1A5276",
                "FFC0392B","FFC0392B","FFC0392B","FF7B241C",
                "FFC0392B","FF4A235A","FF0E7060","FF0E7060","FF7D6608","FF16213E"]

    for ri,((lbl,args),color) in enumerate(zip(sits,sit_colors)):
        r=ri+3; ws7.row_dimensions[r].height=34
        sc(ws7,r,1,lbl,bold=True,sz=9,fc=CW,bg=color)
        sp=sit(**args)
        sr=[p for p in sp if p['rp']=='Run']; spass=[p for p in sp if p['rp']=='Pass']
        blitz_p=[p for p in sp if p['blitz'] not in ('','nan','None','0','No')]
        l_p=[p for p in sp if p['hash']=='L']; m_p=[p for p in sp if p['hash']=='M']; r_p=[p for p in sp if p['hash']=='R']
        l_r=len([p for p in l_p if p['rp']=='Run']); m_r=len([p for p in m_p if p['rp']=='Run']); r_r=len([p for p in r_p if p['rp']=='Run'])
        tf=top3(sp,'front'); tc=top3(sp,'cov'); tp=top3(sp,'play')
        sc(ws7,r,2,len(sr),bold=True,sz=12,fc="FF8B0000",bg="FFFDE8E8",fmt="0")
        sc(ws7,r,3,len(spass),bold=True,sz=12,fc="FF00008B",bg="FFE8F0FE",fmt="0")
        sc(ws7,r,4,tf[0]['v']+f" ({tf[0]['n']})" if tf else "—",sz=9,bg="FFFDE8E8",wrap=True,h="left")
        sc(ws7,r,5,tc[0]['v']+f" ({tc[0]['n']})" if tc else "—",sz=9,bg="FFE8F0FE",wrap=True,h="left")
        sc(ws7,r,6,round(len(blitz_p)/len(sp),2) if sp else "",sz=10,fc="FF4A235A",bg="FFEDE7F6",fmt="0%")
        sc(ws7,r,7,round(l_r/len(l_p),2) if l_p else "",sz=10,fc="FF6C3483",bg="FFEAF0FF",fmt="0%")
        sc(ws7,r,8,round(m_r/len(m_p),2) if m_p else "",sz=10,fc="FF1A5276",bg="FFE8F0FE",fmt="0%")
        sc(ws7,r,9,round(r_r/len(r_p),2) if r_p else "",sz=10,fc="FF784212",bg="FFFFF0EA",fmt="0%")
        sc(ws7,r,10,tp[0]['v']+f" ({tp[0]['n']})" if tp else "—",sz=9,bg=CL,wrap=True,h="left")
        sc(ws7,r,11,"",bg=CL if ri%2==0 else CW,sz=9,wrap=True,v="top")
    ws7.freeze_panes="D3"

    # ── Tab 8: OC Call Sheet ─────────────────────────────────
    ws8=wb2.create_sheet("8. OC Call Sheet Builder")
    ws8.sheet_properties.tabColor="0E7060"; ws8.sheet_view.showGridLines=False
    NC8=11; widths(ws8,[16,10,12,20,20,18,18,18,18,9,28])
    banner(ws8,1,"OC CALL SHEET BUILDER  —  Fill in your calls based on tendencies",NC8,bg=CTe,sz=12,ht=30)
    c8h=["Situation","Field\nZone","Down /\nDist","Expected Front","Expected Coverage",
         "Blitz Alert","Best Run","Best Pass","Best Protection","Priority","Notes"]
    ws8.row_dimensions[2].height=38
    for ci,h in enumerate(c8h): hdr(ws8,2,ci+1,h,bg=CTe,sz=9,wrap=True)
    call_sits=["1st & 10","1st & 10 (Own Half)","1st & 10 (Opp Half)",
        "2nd & Long (8+)","2nd & Medium (4-7)","2nd & Short (1-3)",
        "3rd & Long (7+)","3rd & Medium (4-6)","3rd & Short (1-3)",
        "4th Down","Red Zone — 1st","Red Zone — 2nd","Red Zone — 3rd",
        "Goal Line","Backed Up","Coming Out","Two-Minute (Lead)",
        "Two-Minute (Trail)","Must-Have Plays","Two-Point Play","Overtime"]
    for ri in range(60):
        r=ri+3; ws8.row_dimensions[r].height=20
        bg="FFF0FFF0" if ri%2==0 else CW
        sv=call_sits[ri] if ri<len(call_sits) else ""
        sc(ws8,r,1,sv,bold=bool(sv),sz=9,bg=bg)
        for ci in range(1,NC8): sc(ws8,r,ci+1,"",bg=bg,sz=9,wrap=(ci>=3),v="top")
    ws8.freeze_panes="D3"

    # ── Tab 9: Coordinator Summary ───────────────────────────
    ws9=wb2.create_sheet("9. Coordinator Summary")
    ws9.sheet_properties.tabColor="F1C40F"; ws9.sheet_view.showGridLines=False
    ws9.page_setup.paperSize=1; ws9.page_setup.orientation="landscape"
    ws9.page_setup.fitToPage=True; ws9.page_setup.fitToWidth=1; ws9.page_setup.fitToHeight=1
    widths(ws9,[22,32,3,22,32,3,22,32])
    banner(ws9,1,"OFFENSIVE COORDINATOR SUMMARY  ·  PRINT LANDSCAPE",8,bg=CB,sz=14,ht=36)
    ws9.row_dimensions[2].height=18
    for lbl,ci in[("OPP:",1),("WEEK:",3),("DATE:",5),("OC:",7)]:
        sc(ws9,2,ci,lbl,bold=True,sz=9,fc=CW,bg=CBl,h="right")
        val={"OPP:":opp,"WEEK:":week,"DATE:":date}.get(lbl,"")
        sc(ws9,2,ci+1,val,bg=CL,sz=9)
    for r in range(3,66):
        ws9.row_dimensions[r].height=16
        for dc in[3,6]: ws9.cell(row=r,column=dc).fill=fil("FFCCCCCC")

    rz=[p for p in plays if p['zone']=='RZ']; gl=[p for p in plays if p['zone']=='GL']
    t3d=[p for p in plays if p['dn']==3]
    all_blitz=[p for p in plays if p['blitz'] not in ('','nan','None','0','No')]

    def blk(ws,sr,ca,cb,title,data,hc):
        ws.row_dimensions[sr].height=17
        ws.merge_cells(start_row=sr,start_column=ca,end_row=sr,end_column=cb)
        c=ws.cell(row=sr,column=ca,value=title)
        c.font=Font(name=FN,bold=True,sz=9,color=CW); c.fill=fil(hc)
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bdr()
        for i,(lbl,val) in enumerate(data):
            rr=sr+1+i; bg=CL if i%2==0 else CW; ws.row_dimensions[rr].height=16
            la=ws.cell(row=rr,column=ca,value=lbl)
            la.font=Font(name=FN,bold=True,sz=8,color=CDG)
            la.fill=fil(bg); la.alignment=Alignment(horizontal="right",vertical="center"); la.border=bdr()
            vb=ws.cell(row=rr,column=cb,value=val)
            vb.font=Font(name=FN,sz=9,color="FF000000")
            vb.fill=fil(bg); vb.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True); vb.border=bdr()
            if ca+1<cb:
                ws.merge_cells(start_row=rr,start_column=ca+1,end_row=rr,end_column=cb-1)
                for mc in range(ca+1,cb): ws.cell(row=rr,column=mc).fill=fil(bg)
        return sr+1+len(data)

    r=3
    r=blk(ws9,r,1,2,"SNAP COUNTS",[
        ("Total Plays",total),("Total Runs",len(runs)),("Total Passes",len(passes)),
        ("Run %",f"{pct(len(runs),total)}%"),("Pass %",f"{pct(len(passes),total)}%"),
    ],CB)
    r+=1
    r=blk(ws9,r,1,2,"BLITZ OVERVIEW",[
        ("Total Blitzes",len(all_blitz)),
        ("Blitz %",f"{pct(len(all_blitz),total)}%"),
        ("RZ Blitz %",f"{pct(len([p for p in rz if p['blitz'] not in ('','nan','None','0','No')]),len(rz))}%"),
        ("3rd Blitz %",f"{pct(len([p for p in t3d if p['blitz'] not in ('','nan','None','0','No')]),len(t3d))}%"),
    ],"FF4A235A")
    r+=1
    blk(ws9,r,1,2,"RED ZONE / GL",[
        ("RZ Runs",len([p for p in rz if p['rp']=='Run'])),
        ("RZ Passes",len([p for p in rz if p['rp']=='Pass'])),
        ("GL Runs",len([p for p in gl if p['rp']=='Run'])),
        ("GL Passes",len([p for p in gl if p['rp']=='Pass'])),
    ],CR)

    r=3
    r=blk(ws9,r,4,5,"TOP FRONTS FACED",[(f"#{i+1}",x['v']+f" ({x['n']})") for i,x in enumerate(top3(plays,'front'))],CR)
    r+=1
    r=blk(ws9,r,4,5,"TOP COVERAGES FACED",[(f"#{i+1}",x['v']+f" ({x['n']})") for i,x in enumerate(top3(plays,'cov'))],CBl)
    r+=1
    blk(ws9,r,4,5,"TOP BLITZ TYPES",[(f"#{i+1}",x['v']+f" ({x['n']})") for i,x in enumerate(top3(all_blitz,'blitz'))],CPu)

    r=3
    r=blk(ws9,r,7,8,"BEST RUNS VS THEIR D  (fill in)",[("1.",""),("2.",""),("3.",""),("4.",""),("5.","")],CTe)
    r+=1
    r=blk(ws9,r,7,8,"BEST PASSES VS THEIR D  (fill in)",[("1.",""),("2.",""),("3.",""),("4.",""),("5.","")],CBl)
    r+=1
    r=blk(ws9,r,7,8,"BLITZ BEATERS  (fill in)",[("1.",""),("2.",""),("3.",""),("4.","")],CPu)
    r+=1
    blk(ws9,r,7,8,"MUST-HAVE PLAYS  (fill in)",[("Call 1.",""),("Call 2.",""),("Call 3.",""),("Call 4.","")],CB)

    buf=io.BytesIO(); wb2.save(buf); buf.seek(0)
    return buf.getvalue()

# ── HTML Report ───────────────────────────────────────────────
def build_html(plays, opp, week, date):
    zc={"BZ":"#c0392b","OF":"#1a5276","MF":"#0e7060","FZ":"#7d6608","RZ":"#c0392b","GL":"#4a235a"}
    zn={"BZ":"Backed Up","OF":"Open Field","MF":"Midfield","FZ":"Fringe","RZ":"Red Zone","GL":"Goal Line"}
    zone_list=["BZ","OF","MF","FZ","RZ","GL"]
    total=len(plays); runs=[p for p in plays if p['rp']=='Run']; passes=[p for p in plays if p['rp']=='Pass']
    all_blitz=[p for p in plays if p['blitz'] not in ('','nan','None','0','No')]
    rz=[p for p in plays if p['zone']=='RZ']; gl=[p for p in plays if p['zone']=='GL']

    def tags(items,cls=''):
        if not items: return '<span class="ctag">—</span>'
        return ''.join(f'<span class="ctag {cls}">{x["v"]} ({x["n"]})</span>' for x in items)

    zone_cards=''
    for z in zone_list:
        zp=[p for p in plays if p['zone']==z]
        if not zp: continue
        zr2=[p for p in zp if p['rp']=='Run']; zpas=[p for p in zp if p['rp']=='Pass']
        rp=pct(len(zr2),len(zp)); pp=pct(len(zpas),len(zp))
        zb=[p for p in zp if p['blitz'] not in ('','nan','None','0','No')]
        zone_cards+=f'''<div class="zone-card">
          <div class="zone-hdr" style="background:{zc[z]}20;border-bottom:2px solid {zc[z]}">
            <div><div class="zone-badge" style="color:{zc[z]}">{z}</div><div class="zone-sub">{zn[z]}</div></div>
            <div class="zone-plays">{len(zp)} plays</div>
          </div>
          <div class="zone-body">
            <div class="bar-row">
              <div class="bar-labels"><span style="color:#e8a095">RUN {rp}%</span><span style="color:#93d4f0">PASS {pp}%</span></div>
              <div class="bar-bg"><div class="bar-fill" style="background:#c0392b;width:{rp}%"></div></div>
            </div>
            <div class="zone-tags">
              <div class="tag-lbl">Top Fronts</div>{tags(top3(zp,"front"),"f")}
              <div class="tag-lbl" style="margin-top:6px">Top Coverages</div>{tags(top3(zp,"cov"),"c")}
              <div class="tag-lbl" style="margin-top:6px">Blitz % — {pct(len(zb),len(zp))}%</div>{tags(top3(zb,"blitz"),"b")}
            </div>
          </div>
        </div>'''

    hash_cards=''
    for h,lbl,cls,color in[('L','Left Hash','hl','#b388d4'),('M','Middle','hm','#5dade2'),('R','Right Hash','hr','#e59866')]:
        hp=[p for p in plays if p['hash']==h]
        if not hp: hash_cards+=f'<div class="hash-card {cls}"><div class="hc-title" style="color:{color}">{lbl}</div><p style="color:rgba(240,237,232,.25);text-align:center;font-size:12px">No data</p></div>'; continue
        hr2=[p for p in hp if p['rp']=='Run']; hpass=[p for p in hp if p['rp']=='Pass']
        hb=[p for p in hp if p['blitz'] not in ('','nan','None','0','No')]
        tf=top3(hp,'front')
        hash_cards+=f'''<div class="hash-card {cls}">
          <div class="hc-title" style="color:{color}">{lbl}</div>
          <div class="hbig" style="color:{color}">{len(hp)}</div>
          <div class="hsub">total plays</div>
          <div class="hrp">
            <div class="hrp-item"><div class="hrp-lbl">Run %</div><div class="hrp-val" style="color:#c0392b">{pct(len(hr2),len(hp))}%</div></div>
            <div class="hrp-item"><div class="hrp-lbl">Blitz %</div><div class="hrp-val" style="color:#b388d4">{pct(len(hb),len(hp))}%</div></div>
          </div>
          <div class="htc">Top Front: <span style="color:#d4a017">{tf[0]["v"]+" ("+str(tf[0]["n"])+")" if tf else "—"}</span></div>
        </div>'''

    sit_rows=''
    sits2=[("1ST DOWN",lambda p:p['dn']==1),("2ND & LONG",lambda p:p['dn']==2 and p['dist']>=7),
           ("2ND & MED",lambda p:p['dn']==2 and 4<=p['dist']<=6),("2ND & SHORT",lambda p:p['dn']==2 and p['dist']<=3),
           ("3RD & LONG",lambda p:p['dn']==3 and p['dist']>=7),("3RD & MED",lambda p:p['dn']==3 and 4<=p['dist']<=6),
           ("3RD & SHORT",lambda p:p['dn']==3 and p['dist']<=3),("4TH DOWN",lambda p:p['dn']==4),
           ("RED ZONE",lambda p:p['zone']=='RZ'),("GOAL LINE",lambda p:p['zone']=='GL'),("BACKED UP",lambda p:p['zone']=='BZ')]
    for lbl,fn in sits2:
        sp=[p for p in plays if fn(p)]
        sr2=[p for p in sp if p['rp']=='Run']; spass=[p for p in sp if p['rp']=='Pass']
        sb=[p for p in sp if p['blitz'] not in ('','nan','None','0','No')]
        tf=top3(sp,'front'); tc=top3(sp,'cov')
        sit_rows+=f'''<tr>
          <td class="sit-lbl">{lbl}</td>
          <td style="text-align:center;font-family:Barlow Condensed,sans-serif;font-weight:800;font-size:20px;color:#c0392b">{len(sr2)}</td>
          <td style="text-align:center;font-family:Barlow Condensed,sans-serif;font-weight:800;font-size:20px;color:#5dade2">{len(spass)}</td>
          <td style="text-align:center;font-family:Barlow Condensed,sans-serif;font-weight:700;font-size:16px;color:#b388d4">{str(pct(len(sb),len(sp)))+"%"  if sp else "—"}</td>
          <td style="font-family:Share Tech Mono,monospace;font-size:9px;color:#d4a017">{tf[0]["v"]+" ("+str(tf[0]["n"])+")" if tf else "—"}</td>
          <td style="font-family:Share Tech Mono,monospace;font-size:9px;color:#d4a017">{tc[0]["v"]+" ("+str(tc[0]["n"])+")" if tc else "—"}</td>
        </tr>'''

    def con_rows(items,color,max_n):
        if not items: return '<div style="color:rgba(240,237,232,.3);font-size:12px">No data — tag in Hudl to see trends</div>'
        return ''.join(f'<div style="padding:9px 0;border-bottom:1px solid rgba(240,237,232,.1)"><div style="display:flex;justify-content:space-between"><span style="font-family:Share Tech Mono,monospace;font-size:11px;color:{color}">{x["v"]}</span><span style="font-family:Barlow Condensed,sans-serif;font-weight:700;font-size:20px;color:{color}">{x["n"]}</span></div><div style="background:rgba(240,237,232,.06);height:3px;margin-top:4px"><div style="background:{color};height:3px;width:{round(x["n"]/max_n*100)}%"></div></div></div>' for x in items)

    mf=top3(plays,'front')[0]['n'] if top3(plays,'front') else 1
    mc=top3(plays,'cov')[0]['n'] if top3(plays,'cov') else 1
    mb=top3(all_blitz,'blitz')[0]['n'] if top3(all_blitz,'blitz') else 1

    return f'''<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>OffensiveIQ — {opp}</title>
<link href="https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@700;800;900&family=Barlow:wght@300;400;500&family=Share+Tech+Mono&display=swap" rel="stylesheet">
<style>:root{{--field:#0a1628;--chalk:#f0ede8;--red:#c0392b;--gold:#d4a017;--blue:#1a5276;--mid:#1e2d3d;--line:rgba(240,237,232,0.1);}}*{{box-sizing:border-box;margin:0;padding:0;}}body{{background:var(--field);color:var(--chalk);font-family:Barlow,sans-serif;font-size:14px;line-height:1.5;}}nav{{display:flex;align-items:center;justify-content:space-between;padding:14px 40px;border-bottom:1px solid var(--line);background:rgba(10,22,40,.97);}}.logo{{font-family:Barlow Condensed,sans-serif;font-weight:900;font-size:22px;}}.logo span{{color:#1a5276;}}.wrap{{max-width:1200px;margin:0 auto;padding:40px;}}.eyebrow{{font-family:Share Tech Mono,monospace;font-size:10px;letter-spacing:.2em;color:var(--gold);text-transform:uppercase;margin-bottom:10px;}}.rpt-hdr{{display:flex;justify-content:space-between;align-items:flex-end;margin-bottom:32px;padding-bottom:18px;border-bottom:1px solid var(--line);}}.rpt-title{{font-family:Barlow Condensed,sans-serif;font-weight:900;font-size:36px;text-transform:uppercase;}}.rpt-meta{{font-family:Share Tech Mono,monospace;font-size:10px;color:var(--gold);text-align:right;}}.sum-grid{{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:36px;}}.sum-card{{background:var(--mid);border:1px solid var(--line);padding:16px;}}.sum-lbl{{font-size:9px;font-weight:600;letter-spacing:.14em;text-transform:uppercase;color:rgba(240,237,232,.35);margin-bottom:5px;}}.sum-val{{font-family:Barlow Condensed,sans-serif;font-weight:800;font-size:32px;line-height:1;}}.stitle{{font-family:Barlow Condensed,sans-serif;font-weight:800;font-size:18px;text-transform:uppercase;letter-spacing:.06em;margin-bottom:16px;margin-top:36px;padding-bottom:8px;border-bottom:1px solid var(--line);}}.zone-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;}}.zone-card{{background:var(--mid);border:1px solid var(--line);overflow:hidden;}}.zone-hdr{{padding:10px 14px;display:flex;justify-content:space-between;align-items:center;}}.zone-badge{{font-family:Barlow Condensed,sans-serif;font-weight:900;font-size:16px;}}.zone-sub{{font-size:9px;color:rgba(240,237,232,.35);}}.zone-plays{{font-family:Share Tech Mono,monospace;font-size:9px;color:var(--gold);}}.zone-body{{padding:11px 14px;}}.bar-row{{margin-bottom:8px;}}.bar-labels{{display:flex;justify-content:space-between;font-size:9px;font-weight:600;letter-spacing:.06em;text-transform:uppercase;margin-bottom:3px;}}.bar-bg{{background:rgba(240,237,232,.07);height:5px;position:relative;}}.bar-fill{{height:5px;position:absolute;left:0;top:0;}}.zone-tags{{margin-top:8px;border-top:1px solid var(--line);padding-top:8px;}}.tag-lbl{{font-size:8px;font-weight:600;letter-spacing:.1em;text-transform:uppercase;color:rgba(240,237,232,.25);margin-bottom:3px;}}.ctag{{display:inline-block;background:rgba(240,237,232,.05);border:1px solid rgba(240,237,232,.1);font-family:Share Tech Mono,monospace;font-size:8px;padding:2px 4px;margin:1px 1px 1px 0;color:rgba(240,237,232,.6);}}.ctag.f{{border-color:rgba(192,57,43,.4);color:#e8a095;}}.ctag.c{{border-color:rgba(93,173,226,.35);color:#93d4f0;}}.ctag.b{{border-color:rgba(180,136,212,.4);color:#b388d4;}}.hash-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;}}.hash-card{{background:var(--mid);border:1px solid var(--line);padding:18px;text-align:center;}}.hc-title{{font-family:Barlow Condensed,sans-serif;font-weight:800;font-size:12px;letter-spacing:.14em;text-transform:uppercase;margin-bottom:12px;padding-bottom:8px;border-bottom:1px solid var(--line);}}.hbig{{font-family:Barlow Condensed,sans-serif;font-weight:900;font-size:44px;line-height:1;margin-bottom:2px;}}.hsub{{font-size:9px;color:rgba(240,237,232,.3);margin-bottom:10px;}}.hrp{{display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-bottom:8px;}}.hrp-item{{background:rgba(240,237,232,.04);padding:7px;}}.hrp-lbl{{font-size:8px;letter-spacing:.08em;text-transform:uppercase;color:rgba(240,237,232,.3);}}.hrp-val{{font-family:Barlow Condensed,sans-serif;font-weight:700;font-size:18px;}}.htc{{font-family:Share Tech Mono,monospace;font-size:9px;color:rgba(240,237,232,.3);}}.sit-table{{width:100%;border-collapse:collapse;font-size:12px;}}.sit-table th{{background:var(--field);padding:7px 10px;font-family:Barlow Condensed,sans-serif;font-weight:700;font-size:9px;letter-spacing:.1em;text-transform:uppercase;color:rgba(240,237,232,.4);border:1px solid var(--line);text-align:center;}}.sit-table th:first-child{{text-align:left;}}.sit-table td{{border:1px solid var(--line);padding:8px 12px;}}.sit-table tr:nth-child(odd) td{{background:rgba(240,237,232,.02);}}.sit-table tr:nth-child(even) td{{background:var(--mid);}}.sit-lbl{{font-family:Barlow Condensed,sans-serif;font-weight:700;font-size:12px;white-space:nowrap;}}.con-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:24px;}}</style>
</head><body>
<nav><div class="logo">Offensive<span>IQ</span></div><div style="font-family:Share Tech Mono,monospace;font-size:10px;color:rgba(240,237,232,.4)">DEFENSIVE TENDENCY REPORT</div></nav>
<div class="wrap">
<div class="rpt-hdr"><div><div class="eyebrow">// Offensive Coordinator — Defensive Tendency Report</div><div class="rpt-title">{opp} — Defensive Analysis</div></div><div class="rpt-meta">WEEK {week}{("<br>"+date) if date else ""}<br>{total} PLAYS ANALYZED</div></div>
<div class="sum-grid">
  <div class="sum-card"><div class="sum-lbl">Total Plays</div><div class="sum-val">{total}</div></div>
  <div class="sum-card"><div class="sum-lbl">Their Run %</div><div class="sum-val" style="color:#c0392b">{pct(len(runs),total)}%</div></div>
  <div class="sum-card"><div class="sum-lbl">Their Pass %</div><div class="sum-val" style="color:#5dade2">{pct(len(passes),total)}%</div></div>
  <div class="sum-card"><div class="sum-lbl">Blitz %</div><div class="sum-val" style="color:#b388d4">{pct(len(all_blitz),total)}%</div></div>
  <div class="sum-card"><div class="sum-lbl">RZ Run %</div><div class="sum-val" style="color:#d4a017">{pct(len([p for p in rz if p["rp"]=="Run"]),len(rz))}%</div></div>
</div>
<div class="stitle">Field Zone Breakdown</div><div class="zone-grid">{zone_cards}</div>
<div class="stitle">Hash Tendencies</div><div class="hash-grid">{hash_cards}</div>
<div class="stitle">Situational Summary</div>
<table class="sit-table">
  <tr><th style="text-align:left">Situation</th><th>Runs</th><th>Passes</th><th>Blitz %</th><th style="text-align:left">Top Front</th><th style="text-align:left">Top Coverage</th></tr>
  {sit_rows}
</table>
<div class="stitle">Defensive Tendencies</div>
<div class="con-grid">
  <div><div class="eyebrow" style="margin-bottom:12px">// Top Fronts</div>{con_rows(top3(plays,"front"),"#e8a095",mf)}</div>
  <div><div class="eyebrow" style="margin-bottom:12px">// Top Coverages</div>{con_rows(top3(plays,"cov"),"#93d4f0",mc)}</div>
  <div><div class="eyebrow" style="margin-bottom:12px">// Top Blitz Types</div>{con_rows(top3(all_blitz,"blitz"),"#b388d4",mb)}</div>
</div>
</div></body></html>'''

# ── STREAMLIT UI ──────────────────────────────────────────────
st.markdown('<div class="main-title">Offensive<span style="color:#1a5276">IQ</span></div>', unsafe_allow_html=True)
st.markdown('<div style="font-size:16px;color:rgba(240,237,232,.55);margin-bottom:24px;font-weight:300">Upload your Hudl playlist and see every defensive tendency — fronts, coverages, blitzes — by zone, hash, and situation.</div>', unsafe_allow_html=True)

st.info("💡 Tag DEF FRONT, COVERAGE, and BLITZ columns in Hudl while watching film for the most complete report. The more you tag, the more powerful the analysis.")

st.divider()
col1,col2,col3=st.columns(3)
with col1: opp  = st.text_input("Opponent Name", placeholder="e.g. Lincoln High School")
with col2: week = st.text_input("Week", placeholder="e.g. 3")
with col3: date = st.text_input("Game Date", placeholder="e.g. Sept 5, 2026")

st.markdown("---")
uploaded=st.file_uploader("Upload Hudl Playlist Export (.xlsx)", type=['xlsx','xls'],
                           help="Export your playlist from Hudl as Excel and upload here.")

if uploaded and st.button("⚡ RUN ANALYSIS"):
    with st.spinner("Analyzing defensive tendencies..."):
        try:
            df=pd.read_excel(uploaded)
            plays=load_plays(df)
            if len(plays)==0:
                st.error("No plays found. Make sure your file has a PLAY TYPE column with 'Run' or 'Pass' values.")
            else:
                opp_name=opp or "Opponent"
                prog=st.progress(0,"Reading data...")
                prog.progress(30,"Calculating zone tendencies...")
                excel_bytes=build_excel(plays,opp_name,week,date)
                prog.progress(70,"Building HTML report...")
                html_bytes=build_html(plays,opp_name,week,date).encode('utf-8')
                prog.progress(100,"Complete!")

                runs=[p for p in plays if p['rp']=='Run']; passes=[p for p in plays if p['rp']=='Pass']
                all_blitz=[p for p in plays if p['blitz'] not in ('','nan','None','0','No')]
                rz=[p for p in plays if p['zone']=='RZ']

                st.success(f"✅ Analysis complete — {len(plays)} plays analyzed")
                st.divider()

                m1,m2,m3,m4,m5=st.columns(5)
                m1.metric("Total Plays",   len(plays))
                m2.metric("Run %",         f"{pct(len(runs),len(plays))}%")
                m3.metric("Pass %",        f"{pct(len(passes),len(plays))}%")
                m4.metric("Blitz %",       f"{pct(len(all_blitz),len(plays))}%")
                m5.metric("RZ Run %",      f"{pct(len([p for p in rz if p['rp']=='Run']),len(rz))}%")

                st.divider()
                st.markdown("### Download Your Reports")
                d1,d2=st.columns(2)
                fname=(opp_name+"_" if opp_name else "")+(f"Week{week}_" if week else "")+"OffensiveIQ"
                with d1:
                    st.download_button("📊 Download Excel Workbook",data=excel_bytes,
                        file_name=f"{fname}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                with d2:
                    st.download_button("🌐 Download HTML Report",data=html_bytes,
                        file_name=f"{fname}_Report.html",mime="text/html")

                st.divider()
                st.markdown("### Quick Summary")
                z1,z2,z3=st.columns(3)
                with z1:
                    st.markdown("**Top Fronts Faced**")
                    for x in top3(plays,'front'): st.markdown(f"- {x['v']} ({x['n']} plays)")
                    if not any(p['front'] for p in plays): st.markdown("*Tag DEF FRONT in Hudl*")
                with z2:
                    st.markdown("**Top Coverages**")
                    for x in top3(plays,'cov'): st.markdown(f"- {x['v']} ({x['n']} plays)")
                    if not any(p['cov'] for p in plays): st.markdown("*Tag COVERAGE in Hudl*")
                with z3:
                    st.markdown("**Top Blitz Types**")
                    for x in top3(all_blitz,'blitz'): st.markdown(f"- {x['v']} ({x['n']})")
                    if not all_blitz: st.markdown("*Tag BLITZ in Hudl*")

        except Exception as e:
            st.error(f"Error: {str(e)}")
            st.info("Make sure this is a Hudl playlist export with PLAY TYPE, YARD LN, OFF FORM, DN, DIST, HASH columns.")

st.divider()
st.markdown('<div style="font-family:Share Tech Mono,monospace;font-size:10px;color:rgba(240,237,232,.25);text-align:center;padding:20px 0">© 2026 OFFENSIVEIQ · BUILT FOR OFFENSIVE COORDINATORS</div>',unsafe_allow_html=True)
